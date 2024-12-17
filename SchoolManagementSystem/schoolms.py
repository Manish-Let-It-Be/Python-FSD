# SET 2 => Practical 1
# School Management System having Student Teacher Courses

import random
import string
import os
from tabulate import tabulate  
from openpyxl import Workbook, load_workbook

database = "school_management.xlsx"

class StudentManagementSystem:
    def __init__(self):
        self.students = []
        self.staffs = []
        self.courses = []
        self.load_from_excel()

    # Load data from Excel
    def load_from_excel(self):
        if not os.path.exists(database):
            return
        wb = load_workbook(database)
        self.load_students(wb['Students'])
        self.load_staff(wb['Staff'])
        self.load_courses(wb['Courses'])
        wb.close()

    def load_students(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student = {
                'id': row[0],
                'enrollment': row[1],
                'name': row[2],
                'email': row[3],
                'phone': row[4],
                'address': row[5],
                'course': row[6],
                'stream': row[7]
            }
            self.students.append(student)

    def load_staff(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            staff = {
                'id': row[0],
                'name': row[1],
                'email': row[2],
                'phone': row[3],
                'course': row[4],
                'batch': row[5]
            }
            self.staffs.append(staff)

    def load_courses(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            course = {
                'id': row[0],
                'name': row[1],
                'assigned_staff': row[2]
            }
            self.courses.append(course)

    # Save data to Excel
    def save_to_excel(self):
        wb = Workbook()
        self.save_students(wb.create_sheet(title="Students"))
        self.save_staff(wb.create_sheet(title="Staff"))
        self.save_courses(wb.create_sheet(title="Courses"))
        wb.save(database)
        wb.close()

    def save_students(self, sheet):
        headers = ['ID', 'Enrollment', 'Name', 'Email', 'Phone', 'Address', 'Course', 'Stream']
        sheet.append(headers)
        for student in self.students:
            sheet.append([
                student['id'],
                student['enrollment'],
                student['name'],
                student['email'],
                student['phone'],
                student['address'],
                student['course'],
                student['stream']
            ])

    def save_staff(self, sheet):
        headers = ['ID', 'Name', 'Email', 'Phone', 'Course', 'Batch']
        sheet.append(headers)
        for staff in self.staffs:
            sheet.append([
                staff['id'],
                staff['name'],
                staff['email'],
                staff['phone'],
                staff['course'],
                staff['batch']
            ])

    def save_courses(self, sheet):
        headers = ['ID', 'Name', 'Assigned Staff']
        sheet.append(headers)
        for course in self.courses:
            sheet.append([
                course['id'],
                course['name'],
                course['assigned_staff']
            ])

    # Generate unique IDs and enrollment numbers
    def generate_unique_student_id(self):
        while True:
            student_id = random.randint(1000, 9999)
            if not any(student['id'] == student_id for student in self.students):
                return student_id

    def generate_enrollment_number(self):
        while True:
            enrollment = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not any(student['enrollment'] == enrollment for student in self.students):
                return enrollment

    def generate_unique_staff_id(self):
        while True:
            staff_id = 'S' + str(random.randint(1000, 9999))
            if not any(staff['id'] == staff_id for staff in self.staffs):
                return staff_id

    # Validate email and phone
    def validate_email(self, email):
        return '@' in email and '.' in email and not any(
            student['email'] == email for student in self.students
        ) and not any(staff['email'] == email for staff in self.staffs)

    def validate_phone(self, phone):
        return len(phone) == 10 and phone.isdigit()

    # Student management methods
    def add_student(self):
        print("\nAdding New Student")
        print("-" * 20)

        name = input("Enter student name: ").strip()
        if not name:
            print("Error: Name is required!")
            return

        email = input("Enter student email: ").strip()
        if not email:
            print("Error: Email is required!")
            return
        if not self.validate_email(email):
            print("Error: Invalid or duplicate email!")
            return

        phone = input("Enter student phone: ").strip()
        if not self.validate_phone(phone):
            print("Error: Invalid phone number!")
            return

        address = input("Enter student address: ").strip()
        if not address:
            print("Error: Address is required!")
            return

        print("\nAvailable Courses:")
        available_courses = ["B.Tech", "BBA", "MBA", "MCA"]
        for idx, course in enumerate(available_courses, 1):
            print(f"{idx}. {course}")
        
        try:
            course_choice = int(input("Select course (enter number): "))
            if course_choice < 1 or course_choice > len(available_courses):
                print("Invalid course selection!")
                return
            course = available_courses[course_choice - 1]
        except ValueError:
            print("Invalid input!")
            return

        print("\nAvailable Streams for", course)
        streams = {
            "B.Tech": ["CS", "Mechanical", "Electrical", "Civil", "Chemical"],
            "BBA": ["Finance", "Marketing", "HR"],
            "MBA": ["Regular", "Advanced"],
            "MCA": ["Regular"]
        }
        available_streams = streams[course]
        for idx, stream in enumerate(available_streams, 1):
            print(f"{idx}. {stream}")
        
        try:
            stream_choice = int(input("Select stream (enter number): "))
            if stream_choice < 1 or stream_choice > len(available_streams):
                print("Invalid stream selection!")
                return
            stream = available_streams[stream_choice - 1]
        except ValueError:
            print("Invalid input!")
            return

        student = {
            'id': self.generate_unique_student_id(),
            'enrollment': self.generate_enrollment_number(),
            'name': name,
            'email': email,
            'phone': phone,
            'address': address,
            'course': course,
            'stream': stream
        }

        self.students.append(student)
        self.save_to_excel()
        print(f"\nStudent added successfully! ID: {student['id']}")

    def delete_student(self):
        student_id = input("Enter student ID to delete: ")
        try:
            student_id = int(student_id)
            student = next((s for s in self.students if s['id'] == student_id), None)
            if student:
                self.students.remove(student)
                self.save_to_excel()
                print("Student deleted successfully!")
            else:
                print("Student not found!")
        except ValueError:
            print("Invalid ID format!")

    def display_all_students(self):
        if not self.students:
            print("No students found!")
            return
        
        headers = ['ID', 'Enrollment', 'Name', 'Email', 'Phone', 'Course', 'Stream']
        data = [[s['id'], s['enrollment'], s['name'], s['email'], s['phone'], s['course'], s['stream']] 
                for s in self.students]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def display_student_by_id(self):
        student_id = input("Enter student ID: ")
        try:
            student_id = int(student_id)
            student = next((s for s in self.students if s['id'] == student_id), None)
            
            if student:
                self.display_student(student)
            else:
                print("Student not found!")
        except ValueError:
            print("Invalid ID format!")

    def display_student(self, student):
        headers = ['Field', 'Value']
        data = [
            ['ID', student['id']],
            ['Enrollment', student['enrollment']],
            ['Name', student['name']],
            ['Email', student['email']],
            ['Phone', student['phone']],
            ['Address', student['address']],
            ['Course', student['course']],
            ['Stream', student['stream']]
        ]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def update_student(self):
        student_id = input("Enter student ID to update: ")
        try:
            student_id = int(student_id)
            student = next((s for s in self.students if s['id'] == student_id), None)
            if not student:
                print("Student not found!")
                return
            
            print("\nCurrent Details:")
            self.display_student(student)

            email = input("Enter new email (press enter to skip): ").strip()
            if email and self.validate_email(email):
                student['email'] = email

            phone = input("Enter new phone (press enter to skip): ").strip()
            if phone and self.validate_phone(phone):
                student['phone'] = phone

            address = input("Enter new address (press Enter to skip): ").strip()
            if address:
                student['address'] = address

            self.save_to_excel()
            print("Student updated successfully!")
        except ValueError:
            print("Invalid ID format!")

    # Staff management methods
    def add_staff(self):
        print("\nAdding New Staff")
        print("-" * 20)

        name = input("Enter staff name: ").strip()
        if not name:
            print("Error: Name is required!")
            return

        email = input("Enter staff email: ").strip()
        if not email:
            print("Error: Email is required!")
            return
        if not self.validate_email(email):
            print("Error: Invalid or duplicate email!")
            return

        phone = input("Enter staff phone: ").strip()
        if not self.validate_phone(phone):
            print("Error: Invalid phone number!")
            return

        course = input("Enter staff course: ").strip()
        batch = input("Enter staff batch: ").strip()

        staff = {
            'id': self.generate_unique_staff_id(),
            'name': name,
            'email': email,
            'phone': phone,
            'course': course,
            'batch': batch
        }

        self.staffs.append(staff)
        self.save_to_excel()
        print(f"\nStaff added successfully! ID: {staff['id']}")

    def delete_staff(self):
        staff_id = input("Enter staff ID to delete: ")
        if not staff_id.startswith('S'):
            print("Invalid staff ID format!")
            return
        staff = next((s for s in self.staffs if s['id'] == staff_id), None)
        if staff:
            self.staffs.remove(staff)
            self.save_to_excel()
            print("Staff deleted successfully!")
        else:
            print("Staff not found!")

    def display_all_staffs(self):
        if not self.staffs:
            print("No staff found!")
            return
        
        headers = ['ID', 'Name', 'Email', 'Phone', 'Course', 'Batch']
        data = [[s['id'], s['name'], s['email'], s['phone'], s['course'], s['batch']] 
                for s in self.staffs]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def update_staff(self):
        staff_id = input("Enter staff ID to update: ")
        if not staff_id.startswith('S'):
            print("Invalid staff ID format!")
            return
        staff = next((s for s in self.staffs if s['id'] == staff_id), None)
        if not staff:
            print("Staff not found!")
            return
        
        print("\nCurrent Details:")
        print(staff)

        email = input("Enter new email (press enter to skip): ").strip()
        if email and self.validate_email(email):
            staff['email'] = email

        phone = input("Enter new phone (press enter to skip): ").strip()
        if phone and self.validate_phone(phone):
            staff['phone'] = phone

        course = input("Enter new course (press enter to skip): ").strip()
        if course:
            staff['course'] = course

        batch = input("Enter new batch (press enter to skip): ").strip()
        if batch:
            staff['batch'] = batch

        self.save_to_excel()
        print("Staff updated successfully!")

    # Course management methods
    def add_course(self):
        print("\nAdding New Course")
        print("-" * 20)

        name = input("Enter course name: ").strip()
        assigned_staff = input("Enter assigned staff ID: ").strip()

        course = {
            'id': len(self.courses) + 1,
            'name': name,
            'assigned_staff': assigned_staff
        }

        self.courses.append(course)
        self.save_to_excel()
        print(f"\nCourse added successfully! ID: {course['id']}")

    def delete_course(self):
        course_id = input("Enter course ID to delete: ")
        course = next((c for c in self.courses if c['id'] == int(course_id)), None)
        if course:
            self.courses.remove(course)
            self.save_to_excel()
            print("Course deleted successfully!")
        else:
            print("Course not found!")

    def display_all_courses(self):
        if not self.courses:
            print("No courses found!")
            return
        
        headers = ['ID', 'Name', 'Assigned Staff']
        data = [[c['id'], c['name'], c['assigned_staff']] 
                for c in self.courses]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def update_course(self):
        course_id = input("Enter course ID to update: ")
        course = next((c for c in self.courses if c['id'] == int(course_id)), None)
        if not course:
            print("Course not found!")
            return
        
        print("\nCurrent Details:")
        print(course)

        assigned_staff = input("Enter new assigned staff ID (press enter to skip): ").strip()
        if assigned_staff:
            course['assigned_staff'] = assigned_staff

        self.save_to_excel()
        print("Course updated successfully!")

    # Main menu
    def main_menu(self):
        while True:
            print("\nStudent Management System")
            print("1. Student Options")
            print("2. Staff Options")
            print("3. Course Options")
            print("4. Exit")
            
            choice = input("\nEnter your choice: ")
            
            if choice == '1':
                self.student_options()
            elif choice == '2':
                self.staff_options()
            elif choice == '3':
                self.course_options()
            elif choice == '4':
                print("Thank you for using Student Management System!")
                break
            else:
                print("Invalid choice! Please try again.")

    def student_options(self):
        while True:
            print("\nStudent Options")
            print("a. Add Student")
            print("b. Delete Student by ID")
            print("c. Display All Students")
            print("d. Display Student by ID")
            print("e. Update Student by ID")
            print("f. Back to Main Menu")

            choice = input("\nEnter your choice: ")
            if choice == 'a':
                self.add_student()
            elif choice == 'b':
                self.delete_student()
            elif choice == 'c':
                self.display_all_students()
            elif choice == 'd':
                self.display_student_by_id()
            elif choice == 'e':
                self.update_student()
            elif choice == 'f':
                break
            else:
                print("Invalid choice! Please try again.")

    def staff_options(self):
        while True:
            print("\nStaff Options")
            print("a. Add Staff")
            print("b. Delete Staff by ID")
            print("c. Display All Staffs")
            print("d. Update Staff by ID")
            print("e. Back to Main Menu")

            choice = input("\nEnter your choice: ")
            if choice == 'a':
                self.add_staff()
            elif choice == 'b':
                self.delete_staff()
            elif choice == 'c':
                self.display_all_staffs()
            elif choice == 'd':
                self.update_staff()
            elif choice == 'e':
                break
            else:
                print("Invalid choice! Please try again.")

    def course_options(self):
        while True:
            print("\nCourse Options")
            print("a. Add Course")
            print("b. Delete Course by ID")
            print("c. Display All Courses")
            print("d. Update Course by ID")
            print("e. Back to Main Menu")

            choice = input("\nEnter your choice: ")
            if choice == 'a':
                self.add_course()
            elif choice == 'b':
                self.delete_course()
            elif choice == 'c':
                self.display_all_courses()
            elif choice == 'd':
                self.update_course()
            elif choice == 'e':
                break
            else:
                print("Invalid choice! Please try again.")

def main():
    sms = StudentManagementSystem()
    sms.main_menu()

if __name__ == "__main__":
    main()
